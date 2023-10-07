from fastapi import APIRouter
import openpyxl

router = APIRouter()

@router.get("/save")
def save_sheet():
    print("Hello i am sheet")
    news_data = [
    {
        "row_number": 1,
        "location": "Dhaka, Bangladesh",
        "vehicle": "Bus",
        "time": "2023-09-29 01:23:29",
        "dead": 10,
        "injured": 20,
        "newslink": "https://www.example.com/dhaka-bus-accident/",
    },
    {
        "row_number": 2,
        "location": "New York City, USA",
        "vehicle": "Train",
        "time": "2023-09-29 02:45:17",
        "dead": 5,
        "injured": 15,
        "newslink": "https://www.example.com/new-york-train-accident/",
    },
    {
        "row_number": 3,
        "location": "Tokyo, Japan",
        "vehicle": "Car",
        "time": "2023-09-29 03:12:34",
        "dead": 2,
        "injured": 10,
        "newslink": "https://www.example.com/tokyo-car-accident/",
    },
    ]

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Create a new worksheet in the workbook
    ws = wb.active

    # Write the header row for the worksheet
    ws.cell(row=1, column=1).value = "Location"
    ws.cell(row=1, column=2).value = "Vehicle"
    ws.cell(row=1, column=3).value = "Time"
    ws.cell(row=1, column=4).value = "Dead"
    ws.cell(row=1, column=5).value = "Injured"
    ws.cell(row=1, column=6).value = "Newslink"

    # Loop through the news data and write each row of data to the worksheet
    for news_item in news_data:
        ws.cell(row=news_item["row_number"], column=1).value = news_item["location"]
        ws.cell(row=news_item["row_number"], column=2).value = news_item["vehicle"]
        ws.cell(row=news_item["row_number"], column=3).value = news_item["time"]
        ws.cell(row=news_item["row_number"], column=4).value = news_item["dead"]
        ws.cell(row=news_item["row_number"], column=5).value = news_item["injured"]
        ws.cell(row=news_item["row_number"], column=6).value = news_item["newslink"]

    # Save the Excel workbook
    wb.save("news_data.xlsx")




